#!/usr/bin/env python3
"""Audit Korean administrative deliverables and enforce a five-round clean streak.

The auditor never modifies deliverable artifacts. It writes only the requested
JSON audit report. It scans DOCX, HWPX, XLSX, PDF, and HWP relationships, hashes
every file below the package root except that report, and repeats the full audit.
It cannot attest to a human seal, signature, approval, or submission authority.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import tempfile
import zipfile
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

try:
    import pymupdf as fitz
except ImportError:  # Compatibility with older PyMuPDF releases.
    import fitz  # type: ignore[no-redef]
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.cell.cell import MergedCell


SCAN_SUFFIXES = {".docx", ".hwpx", ".xlsx", ".pdf", ".hwp"}
REQUIRED_CLEAN_STREAK = 5
FORBIDDEN_RGB = {
    "D9EAF7", "EAF2F8", "1F4E78", "EEF5FA", "F4F8FB", "4472C4",
    "365F91", "4F81BD", "0000FF", "002060", "FF0000", "C00000",
}
ALLOWED_GRAY = {"D9D9D9", "E7E6E6", "F2F2F2", "FFFFFF", "000000"}
W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
GRAY_HEADER_RGB = {"D9D9D9", "E7E6E6", "F2F2F2"}
NEUTRAL_RGB = GRAY_HEADER_RGB | {"FFFFFF", "000000", "AUTO", "NONE", "TRANSPARENT"}
COLOR_ATTRS = {"color", "fill", "faceColor", "textColor", "lineColor", "borderColor", "shadeColor"}


@dataclass(frozen=True)
class Issue:
    code: str
    path: str
    detail: str


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def all_tree_files(root: Path, excluded: set[Path]) -> list[Path]:
    """Return every regular package file except explicitly excluded outputs."""
    return sorted(
        path for path in root.rglob("*")
        if path.is_file() and path.resolve() not in excluded
    )


def artifact_files(root: Path, excluded: set[Path]) -> list[Path]:
    return [path for path in all_tree_files(root, excluded) if path.suffix.lower() in SCAN_SUFFIXES]


def tree_hash(files: Iterable[Path], root: Path) -> str:
    digest = hashlib.sha256()
    for path in files:
        rel = path.relative_to(root).as_posix().encode("utf-8")
        digest.update(len(rel).to_bytes(4, "big"))
        digest.update(rel)
        digest.update(bytes.fromhex(sha256(path)))
    return digest.hexdigest().upper()


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def attr_value(element: ET.Element, name: str) -> str | None:
    for key, value in element.attrib.items():
        if local_name(key) == name:
            return value
    return None


def parse_zip_xml(path: Path) -> tuple[dict[str, ET.Element], list[Issue]]:
    issues: list[Issue] = []
    roots: dict[str, ET.Element] = {}
    try:
        with zipfile.ZipFile(path) as archive:
            bad = archive.testzip()
            if bad:
                issues.append(Issue("ZIP_CRC", path.name, bad))
            for name in archive.namelist():
                if name.lower().endswith((".xml", ".rels", ".hpf", ".rdf")):
                    try:
                        roots[name] = ET.fromstring(archive.read(name))
                    except ET.ParseError as exc:
                        issues.append(Issue("XML_PARSE", path.name, f"{name}: {exc}"))
    except (OSError, zipfile.BadZipFile) as exc:
        issues.append(Issue("ZIP_OPEN", path.name, str(exc)))
    return roots, issues


def scan_forbidden_xml(path: Path, roots: dict[str, ET.Element], *, members: set[str] | None = None) -> list[Issue]:
    issues: list[Issue] = []
    hits: set[tuple[str, str]] = set()
    for member, root in roots.items():
        if members is not None and member not in members:
            continue
        for element in root.iter():
            for key, value in element.attrib.items():
                candidate = value.replace("#", "").upper()
                if candidate in FORBIDDEN_RGB:
                    hits.add((member, f"{local_name(key)}={candidate}"))
    for member, detail in sorted(hits):
        issues.append(Issue("FORBIDDEN_COLOR_XML", path.name, f"{member}: {detail}"))
    return issues


def normalized_rgb(value: str | None) -> str | None:
    if value is None:
        return None
    candidate = value.strip().lstrip("#").upper()
    if len(candidate) == 8 and re.fullmatch(r"[0-9A-F]{8}", candidate):
        candidate = candidate[-6:]
    return candidate if re.fullmatch(r"[0-9A-F]{6}|AUTO|NONE|TRANSPARENT", candidate) else None


def xml_color_is_forbidden(value: str | None) -> bool:
    candidate = normalized_rgb(value)
    if candidate is None or candidate in NEUTRAL_RGB:
        return False
    if candidate in FORBIDDEN_RGB:
        return True
    if re.fullmatch(r"[0-9A-F]{6}", candidate):
        return visually_forbidden(tuple(int(candidate[index:index + 2], 16) for index in (0, 2, 4)))
    return False


def paragraph_text(paragraph: ET.Element) -> str:
    return "".join(node.text or "" for node in paragraph.iter() if local_name(node.tag) == "t").strip()


def _docx_active_style_ids(roots: dict[str, ET.Element]) -> set[str]:
    """Resolve styles actually referenced by body/header/footer content."""
    active: set[str] = set()
    for name, root in roots.items():
        if name == "word/document.xml" or name.startswith(("word/header", "word/footer")):
            for node in root.iter():
                if local_name(node.tag) in {"pStyle", "rStyle", "tblStyle"}:
                    value = attr_value(node, "val")
                    if value:
                        active.add(value)
    styles = roots.get("word/styles.xml")
    if styles is None:
        return active
    style_map = {
        attr_value(style, "styleId"): style
        for style in styles if local_name(style.tag) == "style" and attr_value(style, "styleId")
    }
    changed = True
    while changed:
        changed = False
        for style_id in tuple(active):
            style = style_map.get(style_id)
            if style is None:
                continue
            for node in style.iter():
                if local_name(node.tag) in {"basedOn", "link", "next"}:
                    linked = attr_value(node, "val")
                    if linked and linked not in active:
                        active.add(linked)
                        changed = True
    return active


def _docx_active_color_issues(path: Path, roots: dict[str, ET.Element]) -> list[Issue]:
    issues: list[Issue] = []
    members = {
        name for name in roots
        if name == "word/document.xml" or name.startswith(("word/header", "word/footer"))
    }
    issues.extend(scan_forbidden_xml(path, roots, members=members))
    styles = roots.get("word/styles.xml")
    if styles is None:
        return issues
    active = _docx_active_style_ids(roots)
    selected = [node for node in styles if local_name(node.tag) == "docDefaults"]
    selected.extend(
        node for node in styles
        if local_name(node.tag) == "style" and attr_value(node, "styleId") in active
    )
    for style in selected:
        label = attr_value(style, "styleId") or "docDefaults"
        for node in style.iter():
            for key, value in node.attrib.items():
                if local_name(key) in {"val", "fill", "color"} and xml_color_is_forbidden(value):
                    issues.append(Issue("DOCX_ACTIVE_STYLE_COLOR", path.name, f"{label}: {local_name(key)}={value}"))
    return issues


def _docx_cell_gray(cell: ET.Element) -> bool:
    return any(
        normalized_rgb(attr_value(node, "fill")) in GRAY_HEADER_RGB
        for node in cell.iter() if local_name(node.tag) == "shd"
    )


def audit_docx(path: Path) -> list[Issue]:
    roots, issues = parse_zip_xml(path)
    document = roots.get("word/document.xml")
    if document is None:
        return issues + [Issue("DOCX_DOCUMENT_XML", path.name, "word/document.xml missing")]
    issues.extend(_docx_active_color_issues(path, roots))

    paragraphs = [node for node in document.iter() if local_name(node.tag) == "p"]
    attachment_re = re.compile(r"^\s*붙임(?:\s*제?\s*\d+\s*호?)?\s*[.:：]?", re.IGNORECASE)
    for paragraph in paragraphs:
        raw = paragraph_text(paragraph)
        compact = re.sub(r"\s+", "", raw)
        if compact == "문서":
            issues.append(Issue("STANDALONE_DOCUMENT_TITLE", path.name, raw))
        if attachment_re.match(raw) and raw.strip() != "붙임":
            jcs = [attr_value(node, "val") for node in paragraph.iter() if local_name(node.tag) == "jc"]
            if any(value in {"center", "right"} for value in jcs):
                issues.append(Issue("ATTACHMENT_HEADING_ALIGNMENT", path.name, raw[:80]))
            if "\t" in raw or re.search(r"붙임\s{3,}\d", raw):
                issues.append(Issue("ATTACHMENT_MECHANICAL_SPACING", path.name, raw[:80]))

    footer_xml = "".join(
        ET.tostring(root, encoding="unicode")
        for member, root in roots.items() if member.startswith("word/footer")
    )
    if not re.search(r"\bPAGE\b", footer_xml, flags=re.IGNORECASE):
        issues.append(Issue("DOCX_PAGE_FIELD", path.name, "PAGE field missing from footer"))

    is_query = bool(re.search(r"(쿼리|query|mapping)", path.name, flags=re.IGNORECASE))
    for table_index, table in enumerate(node for node in document.iter() if local_name(node.tag) == "tbl"):
        tbl_pr = next((node for node in table if local_name(node.tag) == "tblPr"), None)
        layout = [] if tbl_pr is None else [node for node in tbl_pr if local_name(node.tag) == "tblLayout"]
        if not layout or attr_value(layout[0], "type") != "fixed":
            issues.append(Issue("DOCX_TABLE_FIXED", path.name, f"table {table_index}"))
        grid = next((node for node in table if local_name(node.tag) == "tblGrid"), None)
        try:
            widths = [] if grid is None else [int(attr_value(node, "w") or 0) for node in grid if local_name(node.tag) == "gridCol"]
        except ValueError:
            widths = []
        if not widths or any(width <= 0 for width in widths):
            issues.append(Issue("DOCX_TABLE_GRID", path.name, f"table {table_index}"))
        if is_query and len(widths) == 2 and sum(widths) and widths[0] / sum(widths) > 0.35:
            issues.append(Issue("QUERY_LABEL_WIDTH", path.name, f"table {table_index}: ratio={widths[0] / sum(widths):.3f}"))
        rows = [node for node in table if local_name(node.tag) == "tr"]
        header_cells = [] if not rows else [node for node in rows[0] if local_name(node.tag) == "tc"]
        semantic_header = len(rows) > 1 and bool(header_cells) and all(_docx_cell_gray(cell) for cell in header_cells)
        if semantic_header and not any(local_name(node.tag) == "tblHeader" for node in rows[0].iter()):
            issues.append(Issue("DOCX_TABLE_HEADER_REPEAT", path.name, f"table {table_index}"))
        rows_requiring_no_split = rows[1:] if semantic_header else (rows if is_query and table_index > 0 else [])
        for row_index, row in enumerate(rows_requiring_no_split, start=1 if semantic_header else 0):
            if not any(local_name(node.tag) == "cantSplit" for node in row.iter()):
                issues.append(Issue("DOCX_ROW_CANT_SPLIT", path.name, f"table {table_index} row {row_index}"))
        for cell_index, cell in enumerate(node for node in table.iter() if local_name(node.tag) == "tc"):
            widths_nodes = [node for node in cell.iter() if local_name(node.tag) == "tcW"]
            if not widths_nodes or any((attr_value(node, "type") == "dxa" and int(attr_value(node, "w") or 0) <= 0) for node in widths_nodes):
                issues.append(Issue("DOCX_CELL_WIDTH", path.name, f"table {table_index} cell {cell_index}"))
            for tc_mar in (node for node in cell.iter() if local_name(node.tag) == "tcMar"):
                for side in tc_mar:
                    if local_name(side.tag) in {"start", "end", "left", "right"}:
                        try:
                            value = int(attr_value(side, "w") or 0)
                        except ValueError:
                            value = 0
                        if value > 220:
                            issues.append(Issue("CELL_MARGIN", path.name, f"table {table_index} cell {cell_index}: {value} twip"))
            if _docx_cell_gray(cell):
                for paragraph in (node for node in cell.iter() if local_name(node.tag) == "p"):
                    text = paragraph_text(paragraph)
                    if text and "center" not in [attr_value(node, "val") for node in paragraph.iter() if local_name(node.tag) == "jc"]:
                        issues.append(Issue("SHADED_CELL_ALIGNMENT", path.name, f"table {table_index}: {text[:60]}"))
    return issues


def audit_hwpx(path: Path) -> list[Issue]:
    roots, issues = parse_zip_xml(path)
    if not roots:
        return issues
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            if "mimetype" not in names:
                issues.append(Issue("HWPX_MIMETYPE", path.name, "mimetype missing"))
            if not any(name.lower() == "contents/header.xml" for name in names):
                issues.append(Issue("HWPX_HEADER", path.name, "Contents/header.xml missing"))
            if not any(re.search(r"Contents/section\d+\.xml$", name, re.IGNORECASE) for name in names):
                issues.append(Issue("HWPX_SECTION", path.name, "section XML missing"))
    except zipfile.BadZipFile:
        return issues

    sections = [root for name, root in roots.items() if re.search(r"Contents/section\d+\.xml$", name, re.IGNORECASE)]
    header = next((root for name, root in roots.items() if name.lower() == "contents/header.xml"), None)
    for section_index, section in enumerate(sections):
        for element in section.iter():
            for key, value in element.attrib.items():
                if local_name(key) in COLOR_ATTRS and xml_color_is_forbidden(value):
                    issues.append(Issue("HWPX_LIVE_COLOR", path.name, f"section {section_index}: {local_name(key)}={value}"))

    if header is not None:
        definitions: dict[str, list[ET.Element]] = {}
        by_id: dict[tuple[str, str], ET.Element] = {}
        for element in header.iter():
            tag = local_name(element.tag)
            definitions.setdefault(tag, []).append(element)
            element_id = attr_value(element, "id")
            if element_id is not None:
                by_id[(tag, element_id)] = element
        ref_targets = {
            "borderFillIDRef": "borderFill", "charPrIDRef": "charPr",
            "paraPrIDRef": "paraPr", "styleIDRef": "style",
        }
        pending: list[tuple[str, str]] = []
        for section in sections:
            for element in section.iter():
                for key, value in element.attrib.items():
                    target = ref_targets.get(local_name(key))
                    if target:
                        pending.append((target, value))
        selected: set[int] = set()
        selected_elements: list[ET.Element] = []
        while pending:
            tag, ref = pending.pop()
            candidate = by_id.get((tag, ref))
            if candidate is None and ref.isdigit():
                index = int(ref)
                values = definitions.get(tag, [])
                if 0 <= index < len(values):
                    candidate = values[index]
            if candidate is None or id(candidate) in selected:
                continue
            selected.add(id(candidate))
            selected_elements.append(candidate)
            for descendant in candidate.iter():
                for key, value in descendant.attrib.items():
                    target = ref_targets.get(local_name(key))
                    if target:
                        pending.append((target, value))
        for element in selected_elements:
            label = f"{local_name(element.tag)} {attr_value(element, 'id') or '?'}"
            for descendant in element.iter():
                for key, value in descendant.attrib.items():
                    if local_name(key) in COLOR_ATTRS and xml_color_is_forbidden(value):
                        issues.append(Issue("HWPX_ACTIVE_STYLE_COLOR", path.name, f"{label}: {local_name(key)}={value}"))
    serialized = "".join(ET.tostring(root, encoding="unicode") for root in roots.values())
    if not re.search(r"(numType\s*=\s*[\"']PAGE|>PAGE<|pageNum)", serialized, flags=re.IGNORECASE):
        issues.append(Issue("HWPX_PAGE_FIELD", path.name, "page number field not detected"))
    return issues


def _apply_tint(rgb: str, tint: float) -> str:
    values = [int(rgb[index:index + 2], 16) for index in (0, 2, 4)]
    if tint < 0:
        values = [round(value * (1 + tint)) for value in values]
    elif tint > 0:
        values = [round(value * (1 - tint) + 255 * tint) for value in values]
    return "".join(f"{max(0, min(255, value)):02X}" for value in values)


def _theme_colors(workbook) -> list[str]:
    if not workbook.loaded_theme:
        return []
    try:
        root = ET.fromstring(workbook.loaded_theme)
    except ET.ParseError:
        return []
    scheme = next((node for node in root.iter() if local_name(node.tag) == "clrScheme"), None)
    if scheme is None:
        return []
    result: list[str] = []
    for slot in scheme:
        child = next(iter(slot), None)
        value = None if child is None else (attr_value(child, "lastClr") or attr_value(child, "val"))
        if normalized_rgb(value):
            result.append(normalized_rgb(value) or "")
    return result


def rgb_from_openpyxl(color, theme: list[str] | None = None) -> str | None:
    if color is None:
        return None
    value: str | None = None
    if color.type == "rgb" and color.rgb:
        value = str(color.rgb)[-6:].upper()
    elif color.type == "indexed" and color.indexed is not None:
        index = int(color.indexed)
        if 0 <= index < len(COLOR_INDEX):
            value = str(COLOR_INDEX[index])[-6:].upper()
    elif color.type == "theme" and color.theme is not None and theme:
        index = int(color.theme)
        if 0 <= index < len(theme):
            value = theme[index]
    if value and getattr(color, "tint", 0):
        value = _apply_tint(value, float(color.tint))
    return normalized_rgb(value)


def _xlsx_snapshot(path: Path) -> dict:
    workbook = load_workbook(path, read_only=False, data_only=False)
    snapshot: dict = {"sheet_order": list(workbook.sheetnames), "sheets": {}}
    try:
        for sheet in workbook.worksheets:
            cells = {}
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell) or cell.value is None:
                        continue
                    cells[cell.coordinate] = {"value": cell.value, "data_type": cell.data_type}
            snapshot["sheets"][sheet.title] = {
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "merges": sorted(str(item) for item in sheet.merged_cells.ranges),
                "cells": cells,
            }
    finally:
        workbook.close()
    return snapshot


def _xlsx_baseline_issues(path: Path, baseline: Path | None) -> list[Issue]:
    if baseline is None:
        return []
    if not baseline.exists():
        return [Issue("XLSX_BASELINE_MISSING", path.name, baseline.name)]
    try:
        current = _xlsx_snapshot(path)
        original = _xlsx_snapshot(baseline)
    except Exception as exc:
        return [Issue("XLSX_BASELINE_OPEN", path.name, str(exc))]
    if current == original:
        return []
    issues: list[Issue] = []
    if current["sheet_order"] != original["sheet_order"]:
        issues.append(Issue("XLSX_SHEET_ORDER_CHANGED", path.name, "sheet order differs"))
    for title in sorted(set(current["sheets"]) | set(original["sheets"])):
        if current["sheets"].get(title) != original["sheets"].get(title):
            issues.append(Issue("XLSX_CONTENT_CHANGED", path.name, title))
    return issues


def _xlsx_header_role(value) -> str:
    text = "" if value is None else str(value).strip()
    if re.search(r"(^|\s)(id|no\.?)($|\s)|기술\s*ID|식별|번호|순번", text, re.I):
        return "id"
    if re.search(r"Tier\s*[12](?!\s*3)|대분류|중분류|분야|구분|산업|업종", text, re.I):
        return "category"
    return "text"


def audit_xlsx(path: Path, baseline: Path | None = None) -> list[Issue]:
    roots, issues = parse_zip_xml(path)
    if issues and not roots:
        return issues
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:
        return issues + [Issue("XLSX_OPEN", path.name, str(exc))]
    try:
        theme = _theme_colors(workbook)
        for sheet in workbook.worksheets:
            populated = [cell for row in sheet.iter_rows() for cell in row if cell.value not in (None, "")]
            if not populated:
                continue
            header_index = min(cell.row for cell in populated)
            expected_freeze = f"A{header_index + 1}"
            freeze = None if sheet.freeze_panes is None else str(sheet.freeze_panes)
            if freeze != expected_freeze:
                issues.append(Issue("XLSX_FREEZE_PANES", path.name, f"{sheet.title}: {freeze!r}, expected {expected_freeze}"))
            for cell in populated:
                    font_rgb = rgb_from_openpyxl(cell.font.color, theme)
                    fill_rgb = rgb_from_openpyxl(cell.fill.fgColor, theme)
                    if xml_color_is_forbidden(font_rgb):
                        issues.append(Issue("XLSX_FONT_COLOR", path.name, f"{sheet.title}!{cell.coordinate}: {font_rgb}"))
                    if xml_color_is_forbidden(fill_rgb):
                        issues.append(Issue("XLSX_FILL_COLOR", path.name, f"{sheet.title}!{cell.coordinate}: {fill_rgb}"))
                    if font_rgb not in {None, "000000"}:
                        issues.append(Issue("XLSX_FONT_NOT_BLACK", path.name, f"{sheet.title}!{cell.coordinate}: {font_rgb}"))
            header_row = tuple(sheet[header_index])
            if header_row:
                for cell in header_row:
                    if cell.value in (None, ""):
                        continue
                    fill_rgb = rgb_from_openpyxl(cell.fill.fgColor, theme)
                    if cell.fill.fill_type != "solid" or fill_rgb not in GRAY_HEADER_RGB:
                        issues.append(Issue("XLSX_HEADER_FILL", path.name, f"{sheet.title}!{cell.coordinate}: {fill_rgb}"))
                    if cell.alignment.horizontal != "center" or cell.alignment.vertical != "center":
                        issues.append(Issue("XLSX_HEADER_ALIGNMENT", path.name, f"{sheet.title}!{cell.coordinate}"))
            for column in range(1, sheet.max_column + 1):
                role = _xlsx_header_role(sheet.cell(header_index, column).value)
                expected = "center" if role in {"id", "category"} else "left"
                for row_index in range(header_index + 1, sheet.max_row + 1):
                    cell = sheet.cell(row_index, column)
                    if cell.value not in (None, "") and cell.alignment.horizontal != expected:
                        issues.append(Issue("XLSX_BODY_ALIGNMENT", path.name, f"{sheet.title}!{cell.coordinate}: expected {expected}"))
            title_rows = str(sheet.print_title_rows or "").replace("$", "")
            if title_rows not in {f"{header_index}:{header_index}", f"'{sheet.title}'!{header_index}:{header_index}"}:
                issues.append(Issue("XLSX_PRINT_TITLE", path.name, f"{sheet.title}: {sheet.print_title_rows}"))
            if sheet.page_setup.orientation != "landscape":
                issues.append(Issue("XLSX_ORIENTATION", path.name, sheet.title))
            if sheet.page_setup.fitToWidth != 1:
                issues.append(Issue("XLSX_FIT_TO_WIDTH", path.name, sheet.title))
            if sheet.page_setup.fitToHeight != 0:
                issues.append(Issue("XLSX_FIT_TO_HEIGHT", path.name, f"{sheet.title}: {sheet.page_setup.fitToHeight}"))
            if str(sheet.page_setup.paperSize) not in {"9", "A4", "a4"}:
                issues.append(Issue("XLSX_PAPER_SIZE", path.name, f"{sheet.title}: {sheet.page_setup.paperSize}"))
            footer = sheet.oddFooter.center.text or ""
            if "&P" not in footer and "&[Page]" not in footer:
                issues.append(Issue("XLSX_PAGE_FOOTER", path.name, sheet.title))
            for letter, dimension in sheet.column_dimensions.items():
                if dimension.width and dimension.width > 46:
                    issues.append(Issue("XLSX_COLUMN_WIDTH", path.name, f"{sheet.title}!{letter}={dimension.width}"))
    finally:
        workbook.close()
    issues.extend(_xlsx_baseline_issues(path, baseline))
    return issues


def color_tuple_to_rgb(value) -> tuple[int, int, int] | None:
    if value is None:
        return None
    if isinstance(value, int):
        return ((value >> 16) & 255, (value >> 8) & 255, value & 255)
    if isinstance(value, (tuple, list)) and len(value) >= 3:
        return tuple(int(round(float(channel) * 255)) for channel in value[:3])  # type: ignore[return-value]
    return None


def visually_forbidden(rgb: tuple[int, int, int] | None) -> bool:
    if rgb is None:
        return False
    red, green, blue = rgb
    return (red > 150 and green < 105 and blue < 105) or (blue > 115 and blue > red * 1.22 and blue > green * 1.12)


def has_page_number(text: str, page_number: int | None = None, page_count: int | None = None) -> bool:
    compact = re.sub(r"\s+", " ", text).strip()
    number = r"\d+" if page_number is None else re.escape(str(page_number))
    total = r"\d+" if page_count is None else re.escape(str(page_count))
    patterns = (
        rf"(?:^|\s)-\s*{number}\s*-(?:\s|$)",
        rf"(?:^|\s){number}\s*/\s*{total}(?:\s|$)",
        rf"(?:^|\s){number}\s+(?:of|OF|Of)\s+{total}(?:\s|$)",
        rf"^\s*{number}\s*$",
    )
    return any(re.search(pattern, compact) for pattern in patterns)


def pdf_has_page_number(words: list, width: float, height: float, page_number: int, page_count: int) -> bool:
    footer_words = [word for word in words if word[1] > height * 0.93]
    for word in footer_words:
        center = (word[0] + word[2]) / 2
        if str(word[4]).strip() == str(page_number) and abs(center - width / 2) <= width * 0.15:
            return True
    lines: dict[int, list] = {}
    for word in footer_words:
        lines.setdefault(round(word[1] / 3), []).append(word)
    return any(
        has_page_number(" ".join(str(word[4]) for word in sorted(line, key=lambda item: item[0])), page_number, page_count)
        for line in lines.values()
    )


def audit_pdf(path: Path) -> list[Issue]:
    issues: list[Issue] = []
    try:
        document = fitz.open(path)
    except Exception as exc:
        return [Issue("PDF_OPEN", path.name, str(exc))]
    try:
        if document.page_count == 0:
            return [Issue("PDF_PAGE_COUNT", path.name, "zero pages")]
        sizes = set()
        for page_index, page in enumerate(document):
            sizes.add((round(page.rect.width, 1), round(page.rect.height, 1)))
            words = page.get_text("words")
            drawings = page.get_drawings()
            if len("".join(word[4] for word in words).strip()) < 3 and not drawings:
                issues.append(Issue("PDF_BLANK_PAGE", path.name, str(page_index + 1)))
            for word in words:
                x0, y0, x1, y1 = word[:4]
                if x0 < -1 or y0 < -1 or x1 > page.rect.width + 1 or y1 > page.rect.height + 1:
                    issues.append(Issue("PDF_TEXT_OVERFLOW", path.name, f"page {page_index + 1}: {word[4]}"))
                    break
            if not pdf_has_page_number(words, page.rect.width, page.rect.height, page_index + 1, document.page_count):
                issues.append(Issue("PDF_PAGE_NUMBER", path.name, f"page {page_index + 1}"))
            text_dict = page.get_text("dict")
            bad_text_colors = set()
            for block in text_dict.get("blocks", []):
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        rgb = color_tuple_to_rgb(span.get("color"))
                        if visually_forbidden(rgb):
                            bad_text_colors.add(rgb)
            if bad_text_colors:
                issues.append(Issue("PDF_TEXT_COLOR", path.name, f"page {page_index + 1}: {sorted(bad_text_colors)}"))
            bad_drawing = False
            for drawing in drawings:
                if visually_forbidden(color_tuple_to_rgb(drawing.get("fill"))) or visually_forbidden(color_tuple_to_rgb(drawing.get("color"))):
                    bad_drawing = True
                    break
            if bad_drawing:
                issues.append(Issue("PDF_DRAWING_COLOR", path.name, f"page {page_index + 1}"))
        if len(sizes) > 1:
            issues.append(Issue("PDF_PAGE_SIZE", path.name, str(sorted(sizes))))
        first = re.sub(r"\s+", "", document[0].get_text("text"))
        if first.startswith("문서"):
            issues.append(Issue("PDF_STANDALONE_DOCUMENT_TITLE", path.name, "first page begins with 문서"))
    finally:
        document.close()
    return issues


def complete_audit(root: Path, excluded: set[Path], baseline_root: Path | None = None) -> tuple[list[Issue], list[Path], list[Path], str]:
    files = artifact_files(root, excluded)
    package_files = all_tree_files(root, excluded)
    issues: list[Issue] = []
    if not files:
        issues.append(Issue("NO_ARTIFACTS", ".", "no supported files found"))
    for path in files:
        suffix = path.suffix.lower()
        if suffix == ".docx":
            file_issues = audit_docx(path)
        elif suffix == ".hwpx":
            file_issues = audit_hwpx(path)
        elif suffix == ".xlsx":
            baseline = None if baseline_root is None else baseline_root / path.relative_to(root)
            file_issues = audit_xlsx(path, baseline)
        elif suffix == ".pdf":
            file_issues = audit_pdf(path)
        else:
            pdf = path.with_suffix(".pdf")
            file_issues = [] if pdf.exists() else [Issue("HWP_RENDERED_PDF_MISSING", path.name, pdf.name)]
        rel = path.relative_to(root).as_posix()
        issues.extend(Issue(issue.code, rel, issue.detail) for issue in file_issues)
    return (
        sorted(set(issues), key=lambda item: (item.path, item.code, item.detail)),
        files,
        package_files,
        tree_hash(package_files, root),
    )


def _write_zip(path: Path, members: dict[str, str | bytes]) -> None:
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, value in members.items():
            archive.writestr(name, value)


def _make_xlsx_fixture(path: Path, *, good: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["ID", "설명"])
    sheet.append([1, "본문"])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="D9D9D9" if good else "4472C4")
        cell.font = Font(color="000000" if good else "0000FF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet["A2"].font = Font(color="000000")
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet["B2"].font = Font(color="000000")
    sheet["B2"].alignment = Alignment(horizontal="left")
    if good:
        sheet.freeze_panes = "A2"
        sheet.print_title_rows = "1:1"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.oddFooter.center.text = "&P / &N"
    workbook.save(path)
    workbook.close()


def self_check() -> int:
    checks: dict[str, bool] = {
        "blue_detector": visually_forbidden((30, 90, 180)),
        "black_not_forbidden": not visually_forbidden((0, 0, 0)),
        "page_dash": has_page_number("- 1 -", 1, 3),
        "page_slash": has_page_number("1 / 3", 1, 3),
        "page_of": has_page_number("1 of 3", 1, 3),
        "page_centered_word": pdf_has_page_number([(294.0, 809.0, 300.0, 820.0, "1")], 594.0, 841.0, 1, 3),
    }
    with tempfile.TemporaryDirectory(prefix="admin-doc-audit-") as temporary:
        root = Path(temporary)
        docx = root / "active-style.docx"
        _write_zip(docx, {
            "word/document.xml": (
                f'<w:document xmlns:w="{W_NS}"><w:body><w:p><w:pPr><w:pStyle w:val="Blue"/>'
                '<w:jc w:val="left"/></w:pPr><w:r><w:t>본문</w:t></w:r></w:p>'
                '<w:p><w:pPr><w:jc w:val="center"/></w:pPr><w:r><w:t>붙임   1. 항목</w:t></w:r></w:p>'
                '<w:tbl><w:tr><w:tc><w:tcPr><w:shd w:fill="D9D9D9"/></w:tcPr>'
                '<w:p><w:pPr><w:jc w:val="center"/></w:pPr><w:r><w:t>헤더</w:t></w:r></w:p></w:tc></w:tr>'
                '<w:tr><w:tc><w:p><w:r><w:t>값</w:t></w:r></w:p></w:tc></w:tr></w:tbl>'
                '</w:body></w:document>'
            ),
            "word/styles.xml": (
                f'<w:styles xmlns:w="{W_NS}"><w:style w:type="paragraph" w:styleId="Blue">'
                '<w:rPr><w:color w:val="4472C4"/></w:rPr></w:style></w:styles>'
            ),
            "word/footer1.xml": f'<w:ftr xmlns:w="{W_NS}"><w:p><w:instrText> PAGE </w:instrText></w:p></w:ftr>',
        })
        docx_codes = {issue.code for issue in audit_docx(docx)}
        checks["docx_active_style"] = "DOCX_ACTIVE_STYLE_COLOR" in docx_codes
        checks["docx_table_contract"] = {
            "DOCX_TABLE_FIXED", "DOCX_TABLE_GRID", "DOCX_CELL_WIDTH",
            "DOCX_TABLE_HEADER_REPEAT", "DOCX_ROW_CANT_SPLIT",
        }.issubset(docx_codes)
        checks["docx_attachment_contract"] = {
            "ATTACHMENT_HEADING_ALIGNMENT", "ATTACHMENT_MECHANICAL_SPACING",
        }.issubset(docx_codes)

        hwpx = root / "live-color.hwpx"
        _write_zip(hwpx, {
            "mimetype": "application/hwp+zip",
            "Contents/header.xml": '<hh:head xmlns:hh="urn:hh"><hh:charPr id="1" faceColor="#4472C4"/></hh:head>',
            "Contents/section0.xml": '<hs:sec xmlns:hs="urn:hs"><hs:run faceColor="#4472C4" charPrIDRef="1"/><hs:fieldBegin type="PAGE"/></hs:sec>',
        })
        hwpx_codes = {issue.code for issue in audit_hwpx(hwpx)}
        checks["hwpx_live_color"] = "HWPX_LIVE_COLOR" in hwpx_codes
        checks["hwpx_active_style_color"] = "HWPX_ACTIVE_STYLE_COLOR" in hwpx_codes

        bad_xlsx = root / "bad.xlsx"
        good_xlsx = root / "good.xlsx"
        _make_xlsx_fixture(bad_xlsx, good=False)
        _make_xlsx_fixture(good_xlsx, good=True)
        bad_codes = {issue.code for issue in audit_xlsx(bad_xlsx)}
        checks["xlsx_bad_fixture"] = {
            "XLSX_FREEZE_PANES", "XLSX_HEADER_FILL", "XLSX_FONT_COLOR",
            "XLSX_PRINT_TITLE", "XLSX_ORIENTATION", "XLSX_FIT_TO_WIDTH",
            "XLSX_PAGE_FOOTER",
        }.issubset(bad_codes)
        checks["xlsx_good_fixture"] = not audit_xlsx(good_xlsx)
        checks["xlsx_baseline_equal"] = not _xlsx_baseline_issues(good_xlsx, good_xlsx)
        changed_xlsx = root / "changed.xlsx"
        _make_xlsx_fixture(changed_xlsx, good=True)
        changed_book = load_workbook(changed_xlsx)
        changed_book.active["B2"] = "변경"
        changed_book.save(changed_xlsx)
        changed_book.close()
        checks["xlsx_baseline_detects_change"] = any(
            issue.code == "XLSX_CONTENT_CHANGED" for issue in _xlsx_baseline_issues(changed_xlsx, good_xlsx)
        )

        binary_hwp = root / "legacy.hwp"
        binary_hwp.write_bytes(b"fixture")
        package_issues, _, package_files, before = complete_audit(root, set())
        checks["hwp_requires_pdf"] = any(issue.code == "HWP_RENDERED_PDF_MISSING" for issue in package_issues)
        note = root / "README.md"
        note.write_text("one", encoding="utf-8")
        first = tree_hash(all_tree_files(root, set()), root)
        note.write_text("two", encoding="utf-8")
        second = tree_hash(all_tree_files(root, set()), root)
        checks["tree_hash_all_files"] = first != second and note in all_tree_files(root, set()) and len(package_files) >= 4
        report = root / "audit.json"
        report.write_text("report", encoding="utf-8")
        checks["tree_hash_excludes_only_report"] = report not in all_tree_files(root, {report.resolve()}) and note in all_tree_files(root, {report.resolve()})
        checks["tree_hash_is_sha256"] = len(before) == 64
    payload = {"status": "passed" if all(checks.values()) else "failed", "checks": checks}
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if all(checks.values()) else 1


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("root", nargs="?", type=Path)
    parser.add_argument("--report", type=Path)
    parser.add_argument("--rounds", type=int, default=REQUIRED_CLEAN_STREAK)
    parser.add_argument("--baseline", type=Path, help="optional parallel source tree for XLSX content invariants")
    parser.add_argument("--self-check", action="store_true")
    args = parser.parse_args()
    if args.self_check:
        return self_check()
    if args.root is None:
        parser.error("root is required unless --self-check is used")
    if args.rounds < 1:
        parser.error("--rounds must be positive")
    root = args.root.resolve()
    if not root.is_dir():
        parser.error(f"root is not a directory: {root}")
    report_path = (args.report or root / "administrative_document_audit.json").resolve()
    baseline_root = args.baseline.resolve() if args.baseline else None
    if baseline_root is not None and not baseline_root.is_dir():
        parser.error(f"baseline is not a directory: {baseline_root}")
    excluded = {report_path}
    history = []
    clean_streak = 0
    frozen_hash: str | None = None
    final_issues: list[Issue] = []
    cumulative: set[Issue] = set()
    for round_number in range(1, args.rounds + 1):
        issues, files, package_files, current_hash = complete_audit(root, excluded, baseline_root)
        hash_changed = frozen_hash is not None and current_hash != frozen_hash
        if hash_changed:
            issues.append(Issue("ARTIFACT_HASH_CHANGED", ".", f"{frozen_hash} -> {current_hash}"))
        frozen_hash = current_hash
        issues = sorted(set(issues), key=lambda item: (item.path, item.code, item.detail))
        clean_streak = clean_streak + 1 if not issues else 0
        final_issues = issues
        cumulative.update(issues)
        history.append({
            "round": round_number,
            "artifact_count": len(files),
            "package_file_count": len(package_files),
            "tree_sha256": current_hash,
            "issue_count": len(issues),
            "issues": [asdict(issue) for issue in issues],
            "clean_streak": clean_streak,
        })
    saturated = not final_issues and clean_streak >= REQUIRED_CLEAN_STREAK
    status = "passed_saturated" if saturated else ("incomplete_clean_streak" if not final_issues else "failed")
    cumulative_issues = sorted(cumulative, key=lambda item: (item.path, item.code, item.detail))
    payload = {
        "contract": "korean-administrative-document-audit-v2",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "root": ".",
        "baseline": "provided" if baseline_root else "not_provided",
        "requested_rounds": args.rounds,
        "required_clean_streak": REQUIRED_CLEAN_STREAK,
        "status": status,
        "clean_streak": clean_streak,
        "tree_sha256": history[-1]["tree_sha256"],
        "rounds": history,
        "issue_count": len(final_issues),
        "issues": [asdict(issue) for issue in final_issues],
        "cumulative_issue_count": len(cumulative_issues),
        "cumulative_issues": [asdict(issue) for issue in cumulative_issues],
        "human_actions_not_attested": ["seal", "signature", "final approval", "submission authority"],
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if saturated else (2 if status == "incomplete_clean_streak" else 1)


if __name__ == "__main__":
    raise SystemExit(main())
